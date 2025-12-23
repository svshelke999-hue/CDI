#!/usr/bin/env python3
"""
Create Excel Report from CDI Evaluation Results

This script reads the evaluation JSON files and creates a comprehensive Excel sheet
with detailed results for each chart.
"""

import json
import sys
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any

try:
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: Required packages not installed.")
    print("Please install: pip install pandas openpyxl")
    sys.exit(1)


def load_evaluation_data(evaluation_file: str) -> List[Dict[str, Any]]:
    """Load chart evaluation data from JSON file."""
    with open(evaluation_file, 'r', encoding='utf-8') as f:
        return json.load(f)


def load_metrics_data(metrics_file: str) -> Dict[str, Any]:
    """Load overall metrics data from JSON file."""
    with open(metrics_file, 'r', encoding='utf-8') as f:
        return json.load(f)


def format_list(items: List[str], max_length: int = 200) -> str:
    """Format a list of items into a readable string."""
    if not items:
        return ""
    
    text = "; ".join(items)
    if len(text) > max_length:
        text = text[:max_length] + "..."
    return text


def create_chart_details_sheet(writer, data: List[Dict[str, Any]]):
    """Create detailed sheet with one row per chart."""
    rows = []
    
    for chart_data in data:
        chart_name = chart_data.get('chart_name', 'Unknown')
        processing_success = chart_data.get('processing_success', False)
        overall_score = chart_data.get('overall_score', 0.0)
        processing_time = chart_data.get('processing_time', 0.0)
        
        llm_eval = chart_data.get('llm_evaluation', {})
        coverage_score = llm_eval.get('coverage_score', 0.0)
        quality_score = llm_eval.get('quality_score', 0.0)
        completeness_score = llm_eval.get('completeness_score', 0.0)
        accuracy_score = llm_eval.get('accuracy_score', 0.0)
        detailed_analysis = llm_eval.get('detailed_analysis', '')
        
        matched_improvements = llm_eval.get('matched_improvements', [])
        missed_improvements = llm_eval.get('missed_improvements', [])
        extra_recommendations = llm_eval.get('extra_recommendations', [])
        strengths = llm_eval.get('strengths', [])
        weaknesses = llm_eval.get('weaknesses', [])
        
        improvement_note = chart_data.get('improvement_note', {})
        primary_focus = improvement_note.get('primary_focus', [])
        recommended_improvements = improvement_note.get('recommended_improvements', [])
        
        cdi_recommendations = chart_data.get('cdi_recommendations', [])
        
        # Format CDI recommendations by payer
        payer_recommendations = {}
        for rec in cdi_recommendations:
            payer = rec.get('payer', 'Unknown')
            procedure = rec.get('procedure', 'Unknown')
            decision = rec.get('decision', 'Unknown')
            primary_reasons = rec.get('primary_reasons', [])
            
            if payer not in payer_recommendations:
                payer_recommendations[payer] = []
            
            payer_recommendations[payer].append({
                'procedure': procedure,
                'decision': decision,
                'reasons': primary_reasons
            })
        
        # Format payer recommendations as text
        payer_text = []
        for payer, recs in payer_recommendations.items():
            for rec in recs:
                payer_text.append(f"{payer} - {rec['procedure']}: {rec['decision']}")
        
        row = {
            'Chart Name': chart_name,
            'Processing Success': 'Yes' if processing_success else 'No',
            'Overall Score (%)': round(overall_score, 2),
            'Coverage Score (%)': round(coverage_score, 2),
            'Quality Score (%)': round(quality_score, 2),
            'Completeness Score (%)': round(completeness_score, 2),
            'Accuracy Score (%)': round(accuracy_score, 2),
            'Processing Time (s)': round(processing_time, 3),
            'Matched Improvements Count': len(matched_improvements),
            'Missed Improvements Count': len(missed_improvements),
            'Extra Recommendations Count': len(extra_recommendations),
            'Matched Improvements': format_list(matched_improvements, 500),
            'Missed Improvements': format_list(missed_improvements, 500),
            'Extra Recommendations': format_list(extra_recommendations, 500),
            'Strengths': format_list(strengths, 500),
            'Weaknesses': format_list(weaknesses, 500),
            'Primary Focus Areas': format_list(primary_focus, 300),
            'Expected Improvements': format_list(recommended_improvements, 500),
            'CDI Recommendations (Payer/Procedure/Decision)': format_list(payer_text, 500),
            'Detailed Analysis': detailed_analysis[:1000] + "..." if len(detailed_analysis) > 1000 else detailed_analysis
        }
        rows.append(row)
    
    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name='Chart Details', index=False)


def create_summary_sheet(writer, metrics: Dict[str, Any]):
    """Create summary sheet with overall metrics."""
    summary_data = {
        'Metric': [
            'Total Charts Evaluated',
            'Successful Evaluations',
            'Failed Evaluations',
            'Average Overall Score (%)',
            'Average Coverage Score (%)',
            'Average Quality Score (%)',
            'Average Completeness Score (%)',
            'Average Accuracy Score (%)',
            'Total Expected Improvements',
            'Total Matched Improvements',
            'Total Missed Improvements',
            'Total Extra Recommendations',
            'Improvement Coverage Rate (%)',
            'Average Processing Time (s)',
            'Total Processing Time (s)',
            'Total CDI Processing Cost ($)',
            'Total Evaluation Cost ($)',
            'Total Cost ($)'
        ],
        'Value': [
            metrics.get('total_charts', 0),
            metrics.get('successful_processing', 0),
            metrics.get('failed_processing', 0),
            round(metrics.get('avg_overall_score', 0.0), 2),
            round(metrics.get('avg_coverage_score', 0.0), 2),
            round(metrics.get('avg_quality_score', 0.0), 2),
            round(metrics.get('avg_completeness_score', 0.0), 2),
            round(metrics.get('avg_accuracy_score', 0.0), 2),
            metrics.get('total_expected_improvements', 0),
            metrics.get('total_matched_improvements', 0),
            metrics.get('total_missed_improvements', 0),
            metrics.get('total_extra_recommendations', 0),
            round(metrics.get('improvement_coverage_rate', 0.0), 2),
            round(metrics.get('avg_processing_time', 0.0), 3),
            round(metrics.get('total_processing_time', 0.0), 3),
            round(metrics.get('total_cost_usd', 0.0), 4),
            round(metrics.get('total_evaluation_cost_usd', 0.0), 4),
            round(metrics.get('total_cost_usd', 0.0) + metrics.get('total_evaluation_cost_usd', 0.0), 4)
        ]
    }
    
    df = pd.DataFrame(summary_data)
    df.to_excel(writer, sheet_name='Summary', index=False)


def create_payer_performance_sheet(writer, metrics: Dict[str, Any]):
    """Create payer-specific performance sheet."""
    payer_perf = metrics.get('payer_performance', {})
    
    if not payer_perf:
        return
    
    rows = []
    for payer, perf in payer_perf.items():
        rows.append({
            'Payer': payer,
            'Charts Evaluated': perf.get('charts_evaluated', 0),
            'Avg Coverage (%)': round(perf.get('avg_coverage', 0.0), 2),
            'Avg Quality (%)': round(perf.get('avg_quality', 0.0), 2),
            'Avg Completeness (%)': round(perf.get('avg_completeness', 0.0), 2),
            'Avg Accuracy (%)': round(perf.get('avg_accuracy', 0.0), 2)
        })
    
    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name='Payer Performance', index=False)


def create_cdi_recommendations_sheet(writer, data: List[Dict[str, Any]]):
    """Create detailed sheet with CDI recommendations for each chart."""
    rows = []
    
    for chart_data in data:
        chart_name = chart_data.get('chart_name', 'Unknown')
        cdi_recommendations = chart_data.get('cdi_recommendations', [])
        
        for rec in cdi_recommendations:
            payer = rec.get('payer', 'Unknown')
            procedure = rec.get('procedure', 'Unknown')
            decision = rec.get('decision', 'Unknown')
            primary_reasons = rec.get('primary_reasons', [])
            missing_requirements = rec.get('missing_requirements', [])
            suggestions = rec.get('suggestions', [])
            
            row = {
                'Chart Name': chart_name,
                'Payer': payer,
                'Procedure': procedure,
                'Decision': decision,
                'Primary Reasons': format_list(primary_reasons, 400),
                'Missing Requirements': format_list(missing_requirements, 400),
                'Suggestions': format_list(suggestions, 400)
            }
            rows.append(row)
    
    if rows:
        df = pd.DataFrame(rows)
        df.to_excel(writer, sheet_name='CDI Recommendations', index=False)


def format_excel_file(excel_path: str):
    """Apply formatting to the Excel file."""
    wb = load_workbook(excel_path)
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # Format each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        # Format data rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = border
                if cell.column_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set width with some padding, but cap at 50
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze first row
        ws.freeze_panes = 'A2'
    
    wb.save(excel_path)


def main():
    """Main function to create Excel report."""
    import sys
    # Find the most recent evaluation files
    eval_dir = Path("Evaluation_result")
    
    print(f"Checking directory: {eval_dir.absolute()}")
    print(f"Directory exists: {eval_dir.exists()}")
    
    if not eval_dir.exists():
        print(f"ERROR: Evaluation_result directory not found!")
        print("Please run the evaluation script first.")
        sys.exit(1)
    
    # Find most recent files
    evaluation_files = list(eval_dir.glob("ground_truth_chart_evaluations_*.json"))
    metrics_files = list(eval_dir.glob("ground_truth_metrics_*.json"))
    
    if not evaluation_files:
        print("ERROR: No evaluation files found in Evaluation_result directory!")
        sys.exit(1)
    
    # Get most recent files
    evaluation_file = max(evaluation_files, key=lambda p: p.stat().st_mtime)
    metrics_file = max(metrics_files, key=lambda p: p.stat().st_mtime) if metrics_files else None
    
    print(f"Loading evaluation data from: {evaluation_file.name}")
    chart_data = load_evaluation_data(str(evaluation_file))
    
    metrics_data = {}
    if metrics_file:
        print(f"Loading metrics data from: {metrics_file.name}")
        metrics_data = load_metrics_data(str(metrics_file))
    
    # Create Excel file
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = eval_dir / f"CDI_Evaluation_Report_{timestamp}.xlsx"
    
    print(f"\nCreating Excel report: {excel_path.name}")
    
    with pd.ExcelWriter(str(excel_path), engine='openpyxl') as writer:
        # Create summary sheet
        if metrics_data:
            create_summary_sheet(writer, metrics_data)
            create_payer_performance_sheet(writer, metrics_data)
        
        # Create detailed sheets
        create_chart_details_sheet(writer, chart_data)
        create_cdi_recommendations_sheet(writer, chart_data)
    
    # Format the Excel file
    print("Applying formatting...")
    format_excel_file(str(excel_path))
    
    print(f"\n✅ Excel report created successfully!")
    print(f"   Location: {excel_path}")
    print(f"\nSheets created:")
    print(f"   - Summary: Overall metrics")
    print(f"   - Payer Performance: Payer-specific scores")
    print(f"   - Chart Details: Detailed results for each chart")
    print(f"   - CDI Recommendations: All CDI recommendations by chart")


if __name__ == "__main__":
    main()

